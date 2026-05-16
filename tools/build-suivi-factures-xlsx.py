#!/usr/bin/env python3
"""Génère suivi-factures-restaurants-pro.xlsx (formules Excel FR, openpyxl)."""
from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
TOOLS = Path(__file__).resolve().parent
CLIENTS_JSON = TOOLS / "pro-clients.json"
OUT_DIRS = [ROOT / "outils", ROOT / "a-telecharger"]
BASE_NAME = "suivi-factures-restaurants-pro.xlsx"

MAX_ROW = 120
ALERT_THRESHOLD = 500
DEPOSIT_AMOUNT = 500
LAST_PREFILL_ROW = 13

HEADERS = [
    "Date facture",
    "N° facture",
    "Montant facture (€)",
    "Versement (€)",
    "Solde (€)",
    "Alerte solde",
    "Payé ? (Oui / Non)",
    "Alerte facture > 500 €",
    "Encours impayé (€)",
    "Alerte encours",
    "Message envoyé ?",
    "Date relance",
]

HEADER_FILL = PatternFill("solid", fgColor="3D2E24")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SOLDE_HEADER_FILL = PatternFill("solid", fgColor="1B5E20")
INPUT_FILL = PatternFill("solid", fgColor="FFFDE7")
SOLDE_FILL = PatternFill("solid", fgColor="E8F5E9")
SOLDE_FONT = Font(bold=True)
ALERT_FONT = Font(bold=True, color="B00020")
MONEY_FMT = '#,##0.00'


def load_restaurants() -> list[dict]:
    raw = json.loads(CLIENTS_JSON.read_text(encoding="utf-8"))
    clients = raw.get("clients") or []
    if not clients:
        raise SystemExit("pro-clients.json : clients vide")
    return clients


def pad2(n: int) -> str:
    return f"{n:02d}"


def build_invoice_rows(restaurant: dict) -> list[dict]:
    rows = []
    prefix = restaurant.get("invoicePrefix", "FAC")
    paid_months = set(restaurant.get("paidMonths") or [])
    deposit_months = set(restaurant.get("depositMonths") or restaurant.get("paidMonths") or [])
    for month in range(1, 13):
        paid = month in paid_months
        deposit = DEPOSIT_AMOUNT if month in deposit_months else 0
        rows.append(
            {
                "date": f"2026-{pad2(month)}-01",
                "inv": f"{prefix}-2026-{pad2(month)}",
                "deposit": deposit,
                "paid": "Oui" if paid else "Non",
                "message": "Oui" if paid else "Non",
                "relance": f"2026-{pad2(month)}-15" if paid else "",
            }
        )
    return rows


def solde_formula(row: int, opening: float) -> str:
    """Solde cumulé : ne dépend pas de la cellule E précédente (recalcule si C change)."""
    open_part = f"{opening}+" if opening > 0 else ""
    return f"={open_part}SOMME($D$2:D{row})-SOMME($C$2:C{row})"


def sheet_ref(name: str) -> str:
    return "'" + name.replace("'", "''") + "'"


def style_header_row(ws) -> None:
    solde_cols = {4, 5, 6}
    for col, title in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = HEADER_FONT
        cell.fill = SOLDE_HEADER_FILL if col in solde_cols else HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def build_client_sheet(wb: Workbook, restaurant: dict) -> None:
    name = str(restaurant.get("sheetName") or restaurant["name"])[:31]
    ws = wb.create_sheet(title=name)
    style_header_row(ws)

    opening = float(restaurant.get("openingBalance") or 0)
    prefill = build_invoice_rows(restaurant)

    for col, width in enumerate([12, 18, 11, 14, 12, 28, 10, 26, 14, 26, 14, 12], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    for r in range(2, MAX_ROW + 1):
        data = prefill[r - 2] if r - 2 < len(prefill) else None

        ws.cell(row=r, column=1, value=data["date"] if data else None)
        ws.cell(row=r, column=2, value=data["inv"] if data else None)

        c_cell = ws.cell(row=r, column=3, value=0)
        c_cell.number_format = MONEY_FMT
        c_cell.fill = INPUT_FILL

        d_val = data["deposit"] if data else 0
        d_cell = ws.cell(row=r, column=4, value=d_val)
        d_cell.number_format = MONEY_FMT
        d_cell.fill = INPUT_FILL

        e_cell = ws.cell(row=r, column=5)
        e_cell.value = solde_formula(r, opening)
        e_cell.number_format = MONEY_FMT
        e_cell.fill = SOLDE_FILL
        e_cell.font = SOLDE_FONT

        f_cell = ws.cell(row=r, column=6)
        f_cell.value = (
            f'=SI(E{r}<=0;"⚠ SOLDE ÉPUISÉ — faire verser {DEPOSIT_AMOUNT} €";"")'
        )
        f_cell.font = ALERT_FONT

        ws.cell(row=r, column=7, value=data["paid"] if data else "Non")

        ws.cell(row=r, column=8).value = (
            f'=SI(C{r}>{ALERT_THRESHOLD};"⚠ Facture > {ALERT_THRESHOLD} € — à relancer";"")'
        )
        ws.cell(row=r, column=8).font = ALERT_FONT

        ws.cell(row=r, column=9).value = (
            f"=SOMME($C$2:$C${MAX_ROW})-SOMME.SI($C$2:$C${MAX_ROW};$G$2:$G${MAX_ROW};\"Oui\")"
        )
        ws.cell(row=r, column=9).number_format = MONEY_FMT

        j_cell = ws.cell(row=r, column=10)
        j_cell.value = (
            f'=SI(I{r}>{ALERT_THRESHOLD};"⚠ ENVOYER MESSAGE PAIEMENT";"")'
        )
        j_cell.font = ALERT_FONT

        ws.cell(row=r, column=11, value=data["message"] if data else "Non")
        ws.cell(row=r, column=12, value=data["relance"] if data else "")

    ws.freeze_panes = "A2"


def build_synthese(wb: Workbook, restaurants: list[dict]) -> None:
    ws = wb.create_sheet(title="Synthèse", index=0)
    headers = [
        "Restaurant",
        "Solde crédit actuel (€)",
        "Alerte solde",
        "Encours impayé (€)",
        "Alerte encours",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL

    for idx, restaurant in enumerate(restaurants, start=2):
        ref = sheet_ref(str(restaurant.get("sheetName") or restaurant["name"])[:31])
        ws.cell(row=idx, column=1, value=restaurant["name"])
        ws.cell(row=idx, column=2).value = f"={ref}!E{LAST_PREFILL_ROW}"
        ws.cell(row=idx, column=2).number_format = MONEY_FMT
        ws.cell(row=idx, column=3).value = (
            f'=SI(B{idx}<=0;"⚠ RECHARGER {DEPOSIT_AMOUNT} €";"OK")'
        )
        ws.cell(row=idx, column=3).font = ALERT_FONT
        ws.cell(row=idx, column=4).value = (
            f"=SOMME({ref}!$C$2:$C${MAX_ROW})-SOMME.SI({ref}!$C$2:$C${MAX_ROW};{ref}!$G$2:$G${MAX_ROW};\"Oui\")"
        )
        ws.cell(row=idx, column=5).value = (
            f'=SI(D{idx}>{ALERT_THRESHOLD};"⚠ RELANCER";"OK")'
        )


def build_help(wb: Workbook, restaurants: list[dict]) -> None:
    names = " · ".join(r["name"] for r in restaurants)
    ws = wb.create_sheet(title="Mode emploi")
    lines = [
        f"1. Un onglet par restaurant ({names}).",
        "2. Colonne C (jaune) : montant facture — le solde (E) se recalcule tout seul.",
        f"3. Colonne D (jaune) : versement client ({DEPOSIT_AMOUNT} €).",
        "4. Téléchargez le fichier .xlsx sur casadessert.fr/outils/",
        "5. Ouvrez avec Microsoft Excel (pas Aperçu Mac). Après saisie en C : Entrée.",
        "6. Si rien ne bouge : Fichier → Options → Formules → Calcul automatique.",
    ]
    for i, line in enumerate(lines, start=1):
        ws.cell(row=i, column=1, value=line)
    ws.column_dimensions["A"].width = 72


def main() -> None:
    restaurants = load_restaurants()
    wb = Workbook()
    wb.remove(wb.active)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.calcMode = "auto"

    for restaurant in restaurants:
        build_client_sheet(wb, restaurant)
    build_synthese(wb, restaurants)
    build_help(wb, restaurants)

    for out_dir in OUT_DIRS:
        out_dir.mkdir(parents=True, exist_ok=True)
        path = out_dir / BASE_NAME
        wb.save(path)
        print("Written", path)


if __name__ == "__main__":
    main()
